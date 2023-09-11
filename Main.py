from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QPushButton
from PyQt6.QtCore import QTimer, Qt
import datetime
from PyQt6 import QtWidgets, QtCore
from PyQt6.QtGui import QPixmap
import tkinter.filedialog
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from form.Login.Janela_Login import Classe_Login 
from form.Inicio.Janela_Inicio import Classe_Inicio
from form.Usuario.Janela_Cad_Usuario import Classe_Cad_Usuario
from form.Usuario.Janela_Edit_Usuario import Classe_Edit_Usuario
from form.Banco.banco import Classe_Banco
from form.Cliente.Janela_Cad_Cliente import Classe_Cad_Cliente
from form.Cliente.Janela_Edit_Cliente import Classe_Edit_Cliente
from form.Agendamento.Janela_Agendamento import Classe_Agendamento
from form.Agendamento.Janela_EditAtendimento import Classe_Editar_Agendamento


class Main(QMainWindow):
    def __init__(self, ):
        super(Main, self).__init__()
        self.inicio = Classe_Inicio()
        self.login = Classe_Login()
        self.cad_usuario = Classe_Cad_Usuario()
        self.edit_usuario = Classe_Edit_Usuario()
        self.banco = Classe_Banco()
        self.cad_cliente = Classe_Cad_Cliente()
        self.edit_cliente = Classe_Edit_Cliente()
        self.agendamento = Classe_Agendamento()
        self.editar_agendamento = Classe_Editar_Agendamento()
        self.timer = QTimer(self)

    
    ######## --- BOTÕES QUE CHAMA FUNÇÕES INSERIR --- ########
        self.agendamento.Bt_Salvar_agendamento.clicked.connect(self.inser_atendimento)
        self.inicio.TableWidget_Cliente.doubleClicked.connect(self.adicionar_atendimento)
        self.cad_usuario.bt_SalvarUsuario.clicked.connect(self.inserir_usuarios)
        self.cad_cliente.Bt_Salvar.clicked.connect(self.inserir_clientes)
    ######## --- BOTÕES QUE CHAMA FUNÇÕES EXCLUIR --- ########
        self.inicio.Bt_Excluir_Cliente.clicked.connect(self.excluir_clientes)
        self.inicio.Bt_Remover_Usuario.clicked.connect(self.excluir_usuarios)
        ######## --- BOTÕES QUE CHAMA FUNÇÕES ATUALIZAR --- ########
        self.inicio.Bt_AtualizarAtendimento.clicked.connect(self.listar_atendimento)
        self.edit_usuario.bt_SalvarUsuario.clicked.connect(self.salvar_usuario_editado)
        self.edit_cliente.Bt_Salvar.clicked.connect(self.salvar_cliente_editado)
        self.inicio.Bt_Edit_Atendimento.clicked.connect(self.editar_atendimento)
    ######## --- BOTÕES QUE CHAMA FUNÇÕES PESQUISAR --- ########
        self.inicio.tx_BuscaUsuarios.textChanged.connect(self.pesquisar_usuarios)
        self.inicio.tx_BuscaClientes.textChanged.connect(self.pesquisar_clientes)
        self.inicio.tx_BuscaAtendimento.textChanged.connect(self.pesquisar_atendimento)
        self.inicio.Bt_FtStatus.clicked.connect(self.filtro_status)
        self.inicio.Bt_FtData.clicked.connect(self.filtro_data)
    ######## --- BOTÕES QUE CHAMA FUNÇÕES IMPRIMIR --- ########
        self.inicio.Bt_Imprimir_Atendimento.clicked.connect(self.imprimir_atendimento)
    ######## --- BOTÕES QUE CHAMA FUNÇÕES DE LOGIN E TELA DE INICIO --- ########   
        self.inicio.Bt_Expandir.clicked.connect(self.expaandir_left_menu)
        self.inicio.Bt_Sair.clicked.connect(self.fechar_tela_inicio)
        self.login.Bt_Login.clicked.connect(self.verifica_login)
        self.login.Tx_Senha.returnPressed.connect(self.verifica_login)

        
    ######## --- CHAMA STAKEWIDGETS --- ########
        self.inicio.Bt_Inicio.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.tela_inicio()))
        self.tela_inicio()

    def tela_inicio(self):
        self.inicio.Bt_Inicio.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.inicio.inicio))

    
        self.inicio.Bt_Atendimento.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.tela_atendimento()))
        self.listar_atendimento()
        self.tela_atendimento()

    def tela_atendimento(self):
        self.inicio.Bt_Atendimento.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.inicio.atendimento))

        
        self.inicio.Bt_Clientes.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.tela_clientes()))   
        self.listar_clientes()
        self.tela_clientes()
        
    def tela_clientes(self):
        self.inicio.Bt_Clientes.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.inicio.clientes))


        self.inicio.Bt_Usuarios.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.tela_usuarios()))
        self.listar_usuarios()   
        self.tela_usuarios()

    def tela_usuarios(self):
        self.inicio.Bt_Usuarios.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.inicio.usuarios))


        self.inicio.Bt_Suporte.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.tela_suporte()))
        self.tela_suporte()
        
    def tela_suporte(self):
        self.inicio.Bt_Suporte.clicked.connect(
        lambda: self.inicio.stackedWidget.setCurrentWidget(self.inicio.suporte))

    # Chamar telas
        self.inicio.Bt_Add_Usuario.clicked.connect(self.cham_cad_usuario)
        self.inicio.Bt_Edit_Usuario.clicked.connect(self.editar_usuario)
        self.inicio.Bt_Cad_Cliente.clicked.connect(self.chama_cad_cliente)
        self.inicio.Bt_Edit_Cliente.clicked.connect(self.editar_clientes)
        self.inicio.Bt_Cad_Atendimento.clicked.connect(self.chama_cad_atendimento)

    def chama_cad_atendimento(self):
        self.agendamento.show()   

    def chama_cad_cliente(self):
        self.cad_cliente.show()


    def cham_cad_usuario(self):
        self.cad_usuario.show()


    def fechar_tela_inicio(self):
        self.inicio.close() 
        
    ######## --- OCULTAR BOTÕES --- ########
    def oculta_bt_usuarios(self):
        if self.inicio.Bt_Usuarios.isEnabled():
            self.inicio.Bt_Usuarios.setEnabled(False) 

    def oculta_bt_excluir_cliente(self):
        if self.inicio.Bt_Excluir_Cliente.isEnabled():
            self.inicio.Bt_Excluir_Cliente.setEnabled(False) 

    def oculta_bt_editar_cliente(self):
        if self.inicio.Bt_Edit_Cliente.isEnabled():
            self.inicio.Bt_Edit_Cliente.setEnabled(False) 

    def oculta_bt_edit_atendimento(self):
        if self.inicio.Bt_Edit_Atendimento.isEnabled():
            self.inicio.Bt_Edit_Atendimento.setEnabled(False) 

    
    ######## --- EXPANDIR MENU --- ########
    def expaandir_left_menu(self):
            tamanho = self.inicio.frame_lateral.width()
            if tamanho == 100:
                novo_tamanho = 50
                self.inicio.Bt_Inicio.setText("")
                self.inicio.Bt_Atendimento.setText("")
                self.inicio.Bt_Clientes.setText("")
                self.inicio.Bt_Usuarios.setText("")
                self.inicio.Bt_Sair.setText("")
            else:
                novo_tamanho = 100
                self.inicio.Bt_Inicio.setText('Inicio')
                self.inicio.Bt_Atendimento.setText("Agenda")
                self.inicio.Bt_Clientes.setText("Clientes")
                self.inicio.Bt_Usuarios.setText("Usuários")
                self.inicio.Bt_Sair.setText("Sair")
            self.animacao = QtCore.QPropertyAnimation(self.inicio.frame_lateral,  b"minimumWidth")
            self.animacao.setStartValue(tamanho)
            self.animacao.setEndValue(novo_tamanho)
            self.animacao.setDuration(390)
            self.animacao.start()


    # Verificação de login
    def verifica_login(self):
        try:
            login = self.login.Tx_Usuario.text()
            senha_digitada = self.login.Tx_Senha.text()
            self.banco.conectar()
            self.banco.cursorr.execute(
            "SELECT senha1, nivel_de_acesso, permissao FROM agendamento.usuarios WHERE login ='{}'".format(login))
            dados_usuarios = self.banco.cursorr.fetchall() 
            senha_db = dados_usuarios[0][0]   
            nivel_de_acesso = dados_usuarios[0][1]  
            permissao = dados_usuarios[0][2]
            try:
                if senha_digitada == senha_db and nivel_de_acesso == 'ADMINISTRADOR':
                    self.login.Lb_Info_banco.setText('')
                    self.login.Lb_Info.setText('')
                    self.login.carregar()
                    self.login.hide()           
                    self.inicio.showMaximized() 
                    self.inicio.Lb_UserLogado.setText(login)
                    self.banco.query.close()
                elif senha_digitada == senha_db and nivel_de_acesso == 'USUÁRIO' and permissao == 'TODAS':
                    self.login.Lb_Info_banco.setText('')
                    self.login.Lb_Info.setText('')
                    self.login.carregar()
                    self.login.hide()        
                    self.inicio.showMaximized()
                    self.oculta_bt_usuarios()
                    self.banco.query.close()
                elif senha_digitada == senha_db and nivel_de_acesso == 'USUÁRIO' and permissao == 'VISUALIZAR':
                    self.login.Lb_Info_banco.setText('')
                    self.login.Lb_Info.setText('')
                    self.login.carregar()
                    self.login.hide()        
                    self.inicio.showMaximized()
                    self.oculta_bt_usuarios()
                    self.oculta_bt_edit_atendimento()
                    self.oculta_bt_editar_cliente()
                    self.oculta_bt_excluir_cliente()
                    self.banco.query.close()
                else:
                    self.login.Lb_Info.setText("Dados de login incorretos!")    
                    self.banco.query.close()
            except:
                self.login.Lb_Info.setText("Erro ao conectar ao banco de dados!")           
                self.banco.query.close()
        except:
            self.login.Lb_Info_banco.setText("Dados de login incorretos!")
            self.banco.query.close()


    # Atendimento
    def inser_atendimento(self):
        current_date = datetime.date.today()
        formatted_date = current_date.strftime("%d/%m/%Y")
        status = 'AGUARDANDO'
        nome = self.agendamento.tx_Nome.text()
        cpf = self.agendamento.tx_Cpf.text()
        telefone = self.agendamento.tx_Telefone.text()
        endereco =self.agendamento.tx_Endereco.text()
        departamento = self.agendamento.cb_Departamento.currentText()
        assunto = self.agendamento.tx_Assunto.text()
        observacao = self.agendamento.tx_Observacao.text()
        self.banco.conectar()
        sql = "INSERT INTO agendamento.atendimento (status, nome, cpf, telefone, endereco, departamento, assunto, observacao, data) VALUES (%s,%s, %s, %s, %s,%s,%s,%s,%s)"
        # Tupla de valores a serem inseridos no lugar dos parâmetros (%s)
        valores = (status, nome, cpf, telefone, endereco, departamento, assunto, observacao, formatted_date)
        # Executar a consulta SQL com os valores
        self.banco.cursorr.execute(sql, valores)
        self.banco.query.commit()
        self.banco.query.close()
        self.listar_atendimento()
        self.agendamento.limpar_campos()

    def adicionar_atendimento(self):
        global numero_id
        linha = self.inicio.TableWidget_Atendimento.currentRow()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT idclientes FROM agendamento.clientes")
        dados_lidos = self.banco.cursorr.fetchall()
        valor_id = dados_lidos[linha][0]
        self.banco.cursorr.execute("SELECT * FROM agendamento.clientes WHERE idclientes=" + str(valor_id))
        paciente = self.banco.cursorr.fetchall()
        self.banco.query.close()
        self.agendamento.show()
        numero_id = valor_id
        self.agendamento.tx_Nome.setText(str(paciente[0][1]))
        self.agendamento.tx_Cpf.setText(str(paciente[0][2]))
        self.agendamento.tx_Telefone.setText(str(paciente[0][3]))
        self.agendamento.tx_Endereco.setText(str(paciente[0][4]))
    
    def listar_atendimento(self):
        self.inicio.TableWidget_Atendimento.setColumnWidth(0,30)
        self.inicio.TableWidget_Atendimento.setColumnWidth(2,190)
        self.inicio.TableWidget_Atendimento.setColumnWidth(3,100)
        self.inicio.TableWidget_Atendimento.setColumnWidth(4,100)
        self.inicio.TableWidget_Atendimento.setColumnWidth(5,250)
        self.inicio.TableWidget_Atendimento.setColumnWidth(6,100)
        self.inicio.TableWidget_Atendimento.setColumnWidth(7,300)
        self.inicio.TableWidget_Atendimento.setColumnWidth(8,250)
        self.inicio.TableWidget_Atendimento.verticalHeader().hide()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT * FROM agendamento.atendimento")
        dados_lidosa = self.banco.cursorr.fetchall()
        self.inicio.TableWidget_Atendimento.setRowCount(len(dados_lidosa))
        self.inicio.TableWidget_Atendimento.setColumnCount(10)
        for a in range(0, len(dados_lidosa)):
            for b in range(0, 10):
                self.inicio.TableWidget_Atendimento.setItem(
                    a, b, QtWidgets.QTableWidgetItem(str(dados_lidosa[a][b])))
        self.banco.query.close()
        self.muda_cor()

    def muda_cor(self):    
        for i in range(self.inicio.TableWidget_Atendimento.rowCount()):
            text = self.inicio.TableWidget_Atendimento.item(i, 1).text()
            button = QPushButton(text)
            self.inicio.TableWidget_Atendimento.setCellWidget(i, 1, button) 
            if text == "EM ANDAMENTO":
                button.setStyleSheet("background-color: blue ; border-radius:10px; color: rgb(255, 255, 255); font: 75 10pt 'MS Shell Dlg 2';") 
            elif text == "CONCLUÍDO":
                button.setStyleSheet("background-color: green ; border-radius:10px; color: rgb(255, 255, 255); font: 75 10pt 'MS Shell Dlg 2';")
            elif text == "CANCELADO":
                button.setStyleSheet("background-color: red; border-radius:10px; color: rgb(255, 255, 255); font: 75 10pt 'MS Shell Dlg 2';")
            else:
                button.setStyleSheet("background-color: rgb(108, 108, 108); border-radius:10px; color: rgb(255, 255, 255); font: 75 10pt 'MS Shell Dlg 2';")
        
    
    def editar_atendimento(self):
        linha = self.inicio.TableWidget_Atendimento.currentRow()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT idatendimento FROM agendamento.atendimento")
        dados_lidos = self.banco.cursorr.fetchall()
        valor_id = dados_lidos[linha][0]
        self.banco.cursorr.execute("SELECT * FROM agendamento.atendimento WHERE idatendimento=" + str(valor_id))
        atendimento = self.banco.cursorr.fetchall()
        self.editar_agendamento.show()
        self.id_atendimento = valor_id
        self.editar_agendamento.cb_StatusAtendimento.setCurrentText(str(atendimento[0][1]))
        self.editar_agendamento.tx_Nome.setText(str(atendimento[0][2]))
        self.editar_agendamento.tx_Cpf.setText(str(atendimento[0][3]))
        self.editar_agendamento.tx_Telefone.setText(str(atendimento[0][4]))
        self.editar_agendamento.tx_Endereco.setText(str(atendimento[0][5]))
        self.editar_agendamento.cb_Departamento.setCurrentText(str(atendimento[0][6]))
        self.editar_agendamento.tx_Assunto.setText(str(atendimento[0][7]))
        self.editar_agendamento.tx_Observacao.setText(str(atendimento[0][8]))
 
    
    def excluir_atendimento(self):
        self.banco.conectar()
        linha = self.inicio.TableWidget_Atendimento.currentRow()
        self.inicio.TableWidget_Atendimento.removeRow(linha)
    ###Remover no banco###
        self.banco.cursorr.execute("SELECT idatendimento FROM agendamento.atendimento")
        dados_lidos = self.banco.cursorr.fetchall()
        valor_id = dados_lidos[linha][0]
        self.banco.cursorr.execute("DELETE FROM agendamento.atendimento WHERE idatendimento=" + str(valor_id))
        self.banco.query.commit()
        self.banco.query.close()
        self.listar_atendimento()
    
    def pesquisar_atendimento(self):
        self.banco.conectar()
        self.valor_consulta = self.inicio.tx_BuscaAtendimento.text()
        self.banco.cursorr.execute(f"SELECT * FROM agendamento.atendimento WHERE nome LIKE '%{self.valor_consulta}%' or cpf LIKE '%{self.valor_consulta}%'")
        lista = self.banco.cursorr.fetchall()
        lista = list(lista)
        if not lista:
            return  self.alerta_registro()     
        else:   
            self.inicio.TableWidget_Atendimento.setRowCount(0)
            #primeiro for trás
            for idxLinha, linha in enumerate(lista):
                self.inicio.TableWidget_Atendimento.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.inicio.TableWidget_Atendimento.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()
        self.muda_cor()

    
    def filtro_status(self):
        self.banco.conectar()
        self.valor_consulta = self.inicio.Tx_FiltroStatus.currentText()
        self.banco.cursorr.execute(f"SELECT * FROM agendamento.atendimento WHERE status LIKE '%{self.valor_consulta}%'")
        lista = self.banco.cursorr.fetchall()
        lista = list(lista)
        if not lista:
            return  self.alerta_registro()     
        else:   
            self.inicio.TableWidget_Atendimento.setRowCount(0)
            #primeiro for trás
            for idxLinha, linha in enumerate(lista):
                self.inicio.TableWidget_Atendimento.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.inicio.TableWidget_Atendimento.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()
        self.muda_cor()

    
    def filtro_data(self):
        self.banco.conectar()
        self.valor_consulta = self.inicio.Tx_FiltroData.text()
        self.banco.cursorr.execute(f"SELECT * FROM agendamento.atendimento WHERE data LIKE '%{self.valor_consulta}%'")
        lista = self.banco.cursorr.fetchall()
        lista = list(lista)
        if not lista:
            return  self.alerta_registro()     
        else:   
            self.inicio.TableWidget_Atendimento.setRowCount(0)
            #primeiro for trás
            for idxLinha, linha in enumerate(lista):
                self.inicio.TableWidget_Atendimento.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.inicio.TableWidget_Atendimento.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()
        self.muda_cor()

    
    def imprimir_atendimento(self):
        try:
            data = {}
            headers = []
            for col in range(self.inicio.TableWidget_Atendimento.columnCount()):
                header_text = self.inicio.TableWidget_Atendimento.horizontalHeaderItem(col).text()
                headers.append(header_text)
                data[header_text] = [self.inicio.TableWidget_Atendimento.item(row, col).text() for row in range(self.inicio.TableWidget_Atendimento.rowCount())]
            df = pd.DataFrame.from_dict(data)
            local_salvar = tkinter.filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)  # Adiciona o cabeçalho à planilha
            for row in df.itertuples(index=False):
                sheet.append(row)
            for col_index, column_cells in enumerate(sheet.columns, start=1):
                max_length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = adjusted_width
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(fill_type='solid', fgColor='AAAAAA')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            workbook.save(local_salvar)
            self.alerta_exportar_relatorio()
        except:
            self.alerta_erro_exportar_relatorio()
        
    
    def alerta_exportar_relatorio(self):
        msg = QMessageBox()
        msg.setWindowTitle("Parabéns!")
        msg.setText("Arquivo Salvo Com Sucesso!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()
    
    def alerta_erro_exportar_relatorio(self):
        msg = QMessageBox()
        msg.setWindowTitle("Erro!")
        msg.setText("Erro!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()


    ######## --- FUNÇÕES DA TELA DE CLIENTES--- ########
    def inserir_clientes(self):
        nome = self.cad_cliente.tx_Nome.text()
        cpf = self.cad_cliente.tx_cpf.text()
        telefone = self.cad_cliente.tx_Telefone.text()
        endereco = self.cad_cliente.tx_Endereco.text()
        self.banco.conectar()
        sql = "INSERT INTO agendamento.clientes (nome, cpf, telefone, endereco) VALUES (%s, %s, %s, %s)"
        # Tupla de valores a serem inseridos no lugar dos parâmetros (%s)
        valores = (nome, cpf, telefone, endereco)
        # Executar a consulta SQL com os valores
        self.banco.cursorr.execute(sql, valores)
        self.banco.query.commit()
        self.banco.query.close()
        self.listar_clientes()
        self.cad_cliente.limpar_campos()
    
    def listar_clientes(self):
        self.inicio.TableWidget_Cliente.setColumnWidth(0,30)
        self.inicio.TableWidget_Cliente.setColumnWidth(1,355)
        self.inicio.TableWidget_Cliente.setColumnWidth(4,355)
        self.inicio.TableWidget_Cliente.verticalHeader().hide()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT * FROM agendamento.clientes")
        dados_lidosc = self.banco.cursorr.fetchall()
        self.banco.query.close()
        self.inicio.TableWidget_Cliente.setRowCount(len(dados_lidosc))
        self.inicio.TableWidget_Cliente.setColumnCount(5)
        for a in range(0, len(dados_lidosc)):
            for b in range(0, 5):
                self.inicio.TableWidget_Cliente.setItem(
                    a, b, QtWidgets.QTableWidgetItem(str(dados_lidosc[a][b])))
                
                
    def editar_clientes(self):
        linha = self.inicio.TableWidget_Cliente.currentRow()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT idclientes FROM agendamento.clientes")
        dados_lidos = self.banco.cursorr.fetchall()
        valor_id = dados_lidos[linha][0]
        self.banco.cursorr.execute("SELECT * FROM agendamento.clientes WHERE idclientes=" + str(valor_id))
        cliente = self.banco.cursorr.fetchall()
        self.edit_cliente.show()
        self.id_clientes = valor_id
        self.edit_cliente.tx_Nome.setText(str(cliente[0][1]))
        self.edit_cliente.tx_cpf.setText(str(cliente[0][2]))
        self.edit_cliente.tx_Telefone.setText(str(cliente[0][3]))
        self.edit_cliente.tx_Endereco.setText(str(cliente[0][4]))
    

    def salvar_cliente_editado(self):
        nome = self.edit_cliente.tx_Nome.text()
        cpf = self.edit_cliente.tx_cpf.text()
        telefone = self.edit_cliente.tx_Telefone.text()
        endereco = self.edit_cliente.tx_Endereco.text()
        if nome == "":
            self.alt_cad_em_branco()
        elif cpf == "":
            self.alt_cad_em_branco()
        elif telefone == "":
            self.alt_cad_em_branco()
        elif endereco == "":
            self.alt_cad_em_branco()
        else:
            self.banco.conectar()
            self.banco.cursorr.execute("UPDATE agendamento.clientes SET nome = '{}', cpf = '{}', telefone = '{}', endereco ='{}' WHERE idclientes = {}".format(
                nome, cpf, telefone, endereco, self.id_clientes))
            self.banco.query.commit()
            self.banco.query.close()
            self.alerta_user_edit()
            self.edit_cliente.close()
            self.listar_clientes()

    
    def excluir_clientes(self):
        self.banco.conectar()
        linha = self.inicio.TableWidget_Cliente.currentRow()
        self.inicio.TableWidget_Cliente.removeRow(linha)
    ###Remover no banco###
        self.banco.cursorr.execute("SELECT idclientes FROM agendamento.clientes")
        dados_lidos = self.banco.cursorr.fetchall()
        valor_id = dados_lidos[linha][0]
        self.banco.cursorr.execute("DELETE FROM agendamento.clientes WHERE idclientes =" + str(valor_id))
        self.banco.query.commit()
        self.banco.query.close()
        self.listar_clientes()

    
    def pesquisar_clientes(self):
        self.banco.conectar()
        self.valor_consulta = self.inicio.tx_BuscaClientes.text()
        self.banco.cursorr.execute(f"SELECT * FROM agendamento.clientes WHERE nome LIKE '%{self.valor_consulta}%' or cpf LIKE '%{self.valor_consulta}%'")
        lista = self.banco.cursorr.fetchall()
        lista = list(lista)
        if not lista:
            return  self.alerta_registro()     
        else:   
            self.inicio.TableWidget_Cliente.setRowCount(0)
            #primeiro for trás
            for idxLinha, linha in enumerate(lista):
                self.inicio.TableWidget_Cliente.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.inicio.TableWidget_Cliente.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()


    ######## --- FUNÇÕES DA TELA DE USUÁRIOS--- ########
    def inserir_usuarios(self):
        nome =  self.cad_usuario.tx_nome.text()
        login = self.cad_usuario.tx_Login.text()
        senha1 = self.cad_usuario.tx_Senha.text()
        senha_digitada = self.cad_usuario.tx_Senha.text()
        nivel_acesso = self.cad_usuario.cb_Nivel_Acesso.currentText()
        permissao =  self.cad_usuario.cb_Permissoes.currentText()
        self.banco.conectar()
        if nome == "":
            self.alt_cad_em_branco()
        elif login == "":
                self.alt_cad_em_branco()
        elif senha1 == "":
                self.alt_cad_em_branco()
        elif senha_digitada == "":
                self.alt_cad_em_branco()
        elif nivel_acesso == "SELECIONE":
                self.alt_erro_cad_nivel_acesso()
        elif permissao == "SELECIONE":
                self.alt_erro_cad_permissao()
        elif senha1!= senha_digitada:
                self.alt_senha_diferente()
        else:
            self.banco.cursorr.execute("INSERT INTO agendamento.usuarios (nome,login,senha1,senha2,nivel_de_acesso,permissao) VALUES('"+nome+"','"+
                    login+"','"+senha1+"','"+senha_digitada+"','"+nivel_acesso+"','"+permissao+"')")
            self.banco.query.commit()
            self.banco.query.close()
            self.alt_cadastro_sucesso()
            self.listar_usuarios()

    # ALERTAS #
    def alt_erro_cad_nivel_acesso(self):
        msg = QMessageBox()
        msg.setWindowTitle("Alerta!")
        msg.setText("Selecione um Nível de Acesso do Usuário!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()

    def alt_erro_cad_permissao(self):
        msg = QMessageBox()
        msg.setWindowTitle("Alerta!")
        msg.setText("Selecione as Permissões de Acesso do Usuário!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()

    def alt_cad_em_branco(self):
        msg = QMessageBox()
        msg.setWindowTitle("Alerta!")
        msg.setText("Complete os Campos em Branco!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()

    def alt_senha_diferente(self):
        msg = QMessageBox()
        msg.setWindowTitle("Alerta!")
        msg.setText("As Senhas Estão Diferentes!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()

    def alt_cadastro_sucesso(self):
        msg = QMessageBox()
        msg.setWindowTitle("Parabéns!")
        msg.setText("Usuário Cadastrado Com Sucesso!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()
        self.cad_usuario.tx_nome.clear()
        self.cad_usuario.tx_Login.clear()
        self.cad_usuario.tx_Senha.clear()
        self.cad_usuario.tx_Senha2.clear()
        self.cad_usuario.cb_Nivel_Acesso.setCurrentIndex(0)
        self.cad_usuario.cb_Permissoes.setCurrentIndex(0)
    ##################################################################################  
    
    def listar_usuarios(self):
        self.inicio.TableWidget_Usuario.verticalHeader().hide()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT idusuarios, nome, login, nivel_de_acesso, permissao FROM agendamento.usuarios")
        dados_lidos = self.banco.cursorr.fetchall()
        self.banco.query.close()
        self.inicio.TableWidget_Usuario.setRowCount(len(dados_lidos))
        self.inicio.TableWidget_Usuario.setColumnCount(5)
        for a in range(0, len(dados_lidos)):
            for b in range(0,5):
                self.inicio.TableWidget_Usuario.setItem(
                    a, b, QtWidgets.QTableWidgetItem(str(dados_lidos[a][b])))
                
    
    def editar_usuario(self):
        linha = self.inicio.TableWidget_Usuario.currentRow()
        self.banco.conectar()
        self.banco.cursorr.execute("SELECT idusuarios FROM agendamento.usuarios")
        dados_lidos = self.banco.cursorr.fetchall()
        valor_id = dados_lidos[linha][0]
        self.banco.cursorr.execute("SELECT * FROM agendamento.usuarios WHERE idusuarios=" + str(valor_id))
        cliente = self.banco.cursorr.fetchall()
        self.edit_usuario.show()
        self.id_usuarios = valor_id
        self.edit_usuario.tx_nome.setText(str(cliente[0][1]))
        self.edit_usuario.tx_Login.setText(str(cliente[0][2]))
        self.edit_usuario.cb_Nivel_Acesso.setCurrentText(str(cliente[0][5]))
        self.edit_usuario.cb_Permissoes.setCurrentText(str(cliente[0][6]))

    
    def salvar_usuario_editado(self):
        # Valor digitado no lineEdit
        nome = self.edit_usuario.tx_nome.text()
        login = self.edit_usuario.tx_Login.text()
        senha_login_1 = self.edit_usuario.tx_Senha.text()
        senha_login_2 = self.edit_usuario.tx_Senha2.text()
        nivel_acesso = self.edit_usuario.cb_Nivel_Acesso.currentText()
        permissao = self.edit_usuario.cb_Permissoes.currentText()
        if nome == "":
            self.alt_cad_em_branco()
        elif login == "":
                self.alt_cad_em_branco()
        elif senha_login_1 == "":
                self.alt_cad_em_branco()
        elif senha_login_2 == "":
                self.alt_cad_em_branco()
        elif nivel_acesso == "SELECIONE":
                self.alt_erro_cad_nivel_acesso()
        elif permissao == "SELECIONE":
                self.alt_erro_cad_permissao()
        elif senha_login_1 != senha_login_2:
                self.alt_senha_diferente()
        else:
            self.banco.conectar()
            self.banco.cursorr.execute("UPDATE agendamento.usuarios SET nome = '{}', login = '{}', senha1 = '{}', senha2 ='{}', nivel_de_acesso = '{}', permissao = '{}' WHERE idusuarios = {}".format(
                nome, login, senha_login_1, senha_login_2, nivel_acesso, permissao, self.id_usuarios))
            self.banco.query.commit()
            self.alerta_user_edit()
            self.listar_usuarios()
            self.limpar_edit_user()

    ################# ALERTAS SALVAR USUÁRIOS #################
    def alerta_user_edit(self):
        msg = QMessageBox()
        msg.setWindowTitle("Parabéns!")
        msg.setText("Editado com Sucesso!")
        msg.setIcon(QMessageBox.Icon.Information)
        msg.exec()

    def sair_edit_usuario(self):
        self.limpar_edit_user()
        self.close()

    def limpar_edit_user(self):
        self.edit_usuario.tx_nome.clear()
        self.edit_usuario.tx_Login.clear()
        self.edit_usuario.tx_Senha.clear()
        self.edit_usuario.tx_Senha2.clear()
        self.edit_usuario.cb_Nivel_Acesso.setCurrentIndex(0)
        self.edit_usuario.cb_Permissoes.setCurrentIndex(0)
        self.close()
    #################################################################################################
    
    def excluir_usuarios(self):
        self.banco.conectar()
        linha = self.inicio.TableWidget_Usuario.currentRow()
        self.inicio.TableWidget_Usuario.removeRow(linha)
    ###Remover no banco###
        self.banco.cursorr.execute("SELECT idusuarios FROM agendamento.usuarios")
        dados_lidos = self.banco.cursorr.fetchall()
        valor_id = dados_lidos[linha][0]
        self.banco.cursorr.execute("DELETE FROM agendamento.usuarios WHERE idusuarios =" + str(valor_id))
        self.banco.query.commit()
        self.banco.query.close()
        self.listar_usuarios()
        
    def pesquisar_usuarios(self):
        self.banco.conectar()
        self.valor_consulta = self.inicio.tx_BuscaUsuarios.text()
        self.banco.cursorr.execute(f"SELECT * FROM agendamento.usuarios WHERE nome LIKE '%{self.valor_consulta}%' or login LIKE '%{self.valor_consulta}%'")
        lista = self.banco.cursorr.fetchall()
        lista = list(lista)
        if not lista:
            return  self.alerta_registro()     
        else:   
            self.inicio.TableWidget_Usuario.setRowCount(0)
            #primeiro for trás
            for idxLinha, linha in enumerate(lista):
                self.inicio.TableWidget_Usuario.insertRow(idxLinha)
                for idxColuna, coluna in enumerate(linha):
                    self.inicio.TableWidget_Usuario.setItem(idxLinha, idxColuna, QtWidgets.QTableWidgetItem(str(coluna)))
        self.banco.query.commit()
        self.banco.query.close()


    def alerta_registro(self):
        msg = QMessageBox()
        msg.setWindowTitle("Alerta!")
        msg.setText("Registro Não Encontrado!")
        msg.setIcon(QMessageBox.Icon.Information)   
        msg.exec()


if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    system = Main()
    system.inicio.showMaximized()
    sys.exit(app.exec())